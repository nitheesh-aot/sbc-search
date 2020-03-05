import datetime
import os
from tempfile import NamedTemporaryFile

from openpyxl import Workbook
from flask import Flask, request, jsonify, send_from_directory, abort
from flask_cors import CORS
from sqlalchemy import desc, func
from functools import reduce
from dotenv import load_dotenv
from sqlalchemy.orm.exc import NoResultFound

from search_api.models import (
    Corporation,
    CorpOpState,
    CorpState,
    CorpParty,
    CorpName,
    Address,
    Office,
    OfficeType,
    OfficesHeld,
    OfficerType,
    Event,
    app, db #TODO, move this out of models.py
)
from search_api.constants import ADDITIONAL_COLS_ADDRESS, ADDITIONAL_COLS_ACTIVE

load_dotenv(verbose=True)


def create_app(run_mode=os.getenv('FLASK_ENV', 'production')):
    """Return a configured Flask App using the Factory method."""
    #app = Flask(__name__)
    # TODO: uncomment all the comments in this block.
    #app.config.from_object(config.CONFIGURATION[run_mode])
    CORS(app)

    # Configure Sentry
    if app.config.get('SENTRY_DSN', None):
        sentry_sdk.init(
            dsn=app.config.get('SENTRY_DSN'),
            integrations=[FlaskIntegration()]
        )

    #app.register_blueprint(API_BLUEPRINT)

    # setup_jwt_manager(app, jwt)

    # @app.after_request
    # def add_version(response):  # pylint: disable=unused-variable
    #     version = get_run_version()
    #     response.headers['API'] = f'search_api/{version}'
    #     return response

    # register_shellcontext(app)

    @app.route('/')
    def hello():
        return "Welcome to the director search API."

    @app.route('/corporation/search/')
    def corporation_search():
        args = request.args
        results = _get_corporation_search_results(args)

        # Total number of results
        # This is waaay to expensive.
        # total_results = results.count()

        # Pagination
        page = int(args.get("page")) if "page" in args else 1
        results = results.paginate(int(page), 20, False)

        corporations = []
        for row in results.items:
            result_dict = {}

            result_fields = ['corp_num', 'corp_nme']

            result_dict = {key: getattr(row, key) for key in result_fields}

            corporations.append(result_dict)

        return jsonify({'results': corporations})

    @app.route('/corporation/search/export/')
    def corporation_search_export():

        # Query string arguments
        args = request.args

        # Fetching results
        results = _get_corporation_search_results(args)

        # Exporting to Excel
        wb = Workbook()

        export_dir = "/tmp"
        with NamedTemporaryFile(mode='w+b', dir=export_dir, delete=True) as f:

            sheet = wb.active

            # Sheet headers (first row)
            _ = sheet.cell(column=1, row=1, value="Corporation Id")
            _ = sheet.cell(column=2, row=1, value="Corp Name")
            # _ = sheet.cell(column=3, row=1, value="Transition Date")
            # _ = sheet.cell(column=4, row=1, value="Address")
            # _ = sheet.cell(column=5, row=1, value="Postal Code")
            # _ = sheet.cell(column=6, row=1, value="City")
            # _ = sheet.cell(column=7, row=1, value="Province")

            index = 2
            for row in results:
                # Corporation.corp_num
                _ = sheet.cell(column=1, row=index, value=row[0])
                # CorpName.corp_nme
                _ = sheet.cell(column=2, row=index, value=row[1])
                # Corporation.transition_dt
                # _ = sheet.cell(column=3, row=index, value=row[3])
                # # Address.addr_line_1
                # _ = sheet.cell(column=4, row=index, value=row[4])
                # # Address.postal_cd
                # _ = sheet.cell(column=5, row=index, value=row[5])
                # # Address.city
                # _ = sheet.cell(column=6, row=index, value=row[6])
                # # Address.province
                # _ = sheet.cell(column=7, row=index, value=row[7])

                index += 1

            current_date = datetime.datetime.strftime(datetime.datetime.now(), "%Y-%m-%d %H:%M:%S")
            filename = "Corporation Search Results {date}.xlsx".format(date=current_date)
            full_filename_path = "{dir}/{filename}".format(dir=export_dir, filename=filename)
            wb.save(filename=full_filename_path)

            return send_from_directory(export_dir, filename, as_attachment=True)

    @app.route('/corporation/<id>')
    def corporation(id):

        # TODO: move queries to model class.
        result = (
            Corporation.query
            # .join(CorpState, CorpState.corp_num == Corporation.corp_num)
            # .join(CorpOpState, CorpOpState.state_typ_cd == CorpState.state_typ_cd)
            # .join(Office, Office.corp_num == Corporation.corp_num)
            .add_columns(
                Corporation.corp_num,
                Corporation.transition_dt,
                Corporation.admin_email,
                # Office.mailing_addr_id,
                # Office.office_typ_cd,
                # CorpOpState.state_typ_cd,
                # CorpOpState.full_desc
            )
            # .filter(Office.end_event_id == None)
            # .filter(CorpState.end_event_id == None)
            .filter(Corporation.corp_num == id).one())

        corp = result[0]
        offices = Office.query.filter_by(corp_num=id, end_event_id=None)
        names = CorpName.query.filter_by(corp_num=id).order_by(desc(CorpName.end_event_id))

        output = {}
        # TODO: switch to marshmallow.
        output['corp_num'] = corp.corp_num
        output['transition_dt'] = corp.transition_dt
        output['offices'] = []
        for office in offices:
            output['offices'].append({
                'delivery_addr': _normalize_addr(office.delivery_addr_id),  # TODO: get full address.
                'mailing_addr': _normalize_addr(office.mailing_addr_id),
                'office_typ_cd': _format_office_typ_cd(office.office_typ_cd),
                'email_address': office.email_address
            })

        output['admin_email'] = result[3]
        # output['state_typ_cd'] = result[4]
        # output['full_desc'] = result[5]

        output['NAMES'] = []
        for row in names:
            output['NAMES'].append({
                'name': row.corp_nme
            })

        return jsonify(output)

    @app.route('/person/search/')
    def corpparty_search():
        args = request.args
        results = _get_corpparty_search_results(args)

        # Total number of results
        # This is waaay to expensive on a large db.
        # total_results = results.count()

        # Pagination
        page = int(args.get("page")) if "page" in args else 1
        results = results.paginate(int(page), 20, False)

        corp_parties = []
        for row in results.items:
            result_fields = [
                'corp_party_id', 'first_nme', 'middle_nme', 'last_nme', 'appointment_dt', 'cessation_dt',
                'corp_num', 'party_typ_cd']

            result_dict = {key: getattr(row, key) for key in result_fields}
            result_dict['corp_party_id'] = int(result_dict['corp_party_id'])

            _add_additional_cols_to_search_results(args, row, result_dict)

            corp_parties.append(result_dict)

        return jsonify({'results': corp_parties})

    @app.route('/person/search/export/')
    def corpparty_search_export():

        # Query string arguments
        args = request.args

        # Fetching results
        results = _get_corpparty_search_results(args)

        # Exporting to Excel
        wb = Workbook()

        export_dir = "/tmp"
        with NamedTemporaryFile(mode='w+b', dir=export_dir, delete=True) as f:

            sheet = wb.active

            # Sheet headers (first row)
            _ = sheet.cell(column=1, row=1, value="Person Id")
            _ = sheet.cell(column=2, row=1, value="First Name")
            _ = sheet.cell(column=3, row=1, value="Middle Name")
            _ = sheet.cell(column=4, row=1, value="Last Name")
            _ = sheet.cell(column=5, row=1, value="Appointment Date")
            _ = sheet.cell(column=6, row=1, value="Cessation Date")
            _ = sheet.cell(column=7, row=1, value="Corporation Id")
            _ = sheet.cell(column=8, row=1, value="Address")

            index = 2
            for row in results:

                # CorpParty.corp_party_id
                _ = sheet.cell(column=1, row=index, value=row[1])
                # CorpParty.first_nme
                _ = sheet.cell(column=2, row=index, value=row[2])
                # CorpParty.middle_nme
                _ = sheet.cell(column=3, row=index, value=row[3])
                # CorpParty.last_nme
                _ = sheet.cell(column=4, row=index, value=row[4])
                # CorpParty.appointment_dt
                _ = sheet.cell(column=5, row=index, value=row[5])
                # CorpParty.cessation_dt
                _ = sheet.cell(column=6, row=index, value=row[6])
                # Corporation.corp_num
                _ = sheet.cell(column=7, row=index, value=row[7])
                # Address.addr_line_1, Address.addr_line_2, Address.addr_line_3
                _ = sheet.cell(column=8, row=index, value=_merge_corpparty_search_addr_fields(row))

                index += 1

            current_date = datetime.datetime.strftime(datetime.datetime.now(), "%Y-%m-%d %H:%M:%S")
            filename = "Director Search Results {date}.xlsx".format(date=current_date)
            full_filename_path = "{dir}/{filename}".format(dir=export_dir, filename=filename)
            wb.save(filename=full_filename_path)

            return send_from_directory(export_dir, filename, as_attachment=True)

    @app.route('/person/<id>')
    def person(id):
        #try:

        result = (CorpParty.query
            .join(Corporation, Corporation.corp_num == CorpParty.corp_num)
            .add_columns(\
                # CorpParty.corp_party_id,
                # CorpParty.first_nme,
                # CorpParty.middle_nme,
                # CorpParty.last_nme,
                # CorpParty.appointment_dt,
                # CorpParty.cessation_dt,
                # CorpParty.corp_num,
                # CorpParty.delivery_addr_id,
                # CorpParty.party_typ_cd
                Corporation.corp_typ_cd,
                Corporation.admin_email,
                # CorpOpState.state_typ_cd,
                # CorpOpState.full_desc,
            ).filter(CorpParty.corp_party_id==int(id))).one()

        person = result[0]
        result_dict = {}
        name = CorpName.query.filter(CorpName.corp_num == person.corp_num).add_columns(CorpName.corp_nme).filter()[0] #TODO: handle multiple names
        offices = Office.query.filter(Office.corp_num == person.corp_num).all()
        delivery_addr = _normalize_addr(person.delivery_addr_id)
        mailing_addr = _normalize_addr(person.mailing_addr_id)

        states = CorpState.query.filter(
            CorpState.corp_num == person.corp_num,
            CorpState.end_event_id == None).all()

        # TODO : list all, or just the one from the correct time.
        corp_delivery_addr = _normalize_addr(offices[0].delivery_addr_id) if offices else ''
        corp_mailing_addr = _normalize_addr(offices[0].mailing_addr_id) if offices else ''

        # TODO: switch to marshmallow.
        result_dict['corp_party_id'] = int(person.corp_party_id)
        result_dict['first_nme'] = person.first_nme
        result_dict['middle_nme'] = person.middle_nme
        result_dict['last_nme'] = person.last_nme
        result_dict['appointment_dt'] = person.appointment_dt
        result_dict['cessation_dt'] = person.cessation_dt
        result_dict['corp_num'] = person.corp_num
        result_dict['corp_nme'] = name.corp_nme
        result_dict['party_typ_cd'] = person.party_typ_cd
        result_dict['corp_party_email'] = person.email_address
        result_dict['delivery_addr'] = delivery_addr
        result_dict['mailing_addr'] = mailing_addr
        result_dict['corp_delivery_addr'] = corp_delivery_addr
        result_dict['corp_mailing_addr'] = corp_mailing_addr
        result_dict['corp_typ_cd'] = result[1]
        result_dict['corp_admin_email'] = result[2]

        result_dict['states'] = [s.as_dict() for s in states]
        # result_dict['full_desc'] = results[0][14]

        return jsonify(result_dict)

    def _normalize_addr(id):
        if not id:
            return ''

        address = Address.query.filter(Address.addr_id == id).add_columns(
            Address.addr_line_1,
            Address.addr_line_2,
            Address.addr_line_3,
            Address.postal_cd,
            Address.city,
            Address.province,
            Address.country_typ_cd,
            ).one()[0]

        def fn(accumulator, s):
            if s:
                return (accumulator or '') + ', ' + (s or '')
            else:
                return accumulator or ''

        return reduce(fn, [address.addr_line_1, address.addr_line_2, address.addr_line_3, address.city, address.province, address.country_typ_cd])

    def _format_office_typ_cd(office_typ_cd):
        if office_typ_cd == "RG":
            return "Registered"
        elif office_typ_cd == "RC":
            return "Records"

    @app.route('/person/officesheld/<corppartyid>')
    def officesheld(corppartyid):
        results = (OfficerType.query
                .join(OfficesHeld, OfficerType.officer_typ_cd==OfficesHeld.officer_typ_cd)
                .join(CorpParty, OfficesHeld.corp_party_id == CorpParty.corp_party_id)
                #.join(Address, CorpParty.mailing_addr_id == Address.addr_id)
                .join(Event, Event.event_id == CorpParty.start_event_id)
                .add_columns(
                    CorpParty.corp_party_id,
                    OfficerType.officer_typ_cd,
                    OfficerType.short_desc,
                    CorpParty.appointment_dt,
                    Event.event_timestmp
                )
                .filter(CorpParty.corp_party_id==int(corppartyid))
            )

        offices = []
        for row in results:
            result_dict = {}

            result_dict['corp_party_id'] = int(row[1])
            result_dict['officer_typ_cd'] = row[2]
            result_dict['short_desc'] = row[3]
            result_dict['appointment_dt'] = row[4]
            result_dict['year'] = row[5].year

            offices.append(result_dict)


        person = CorpParty.query.filter(CorpParty.corp_party_id==int(corppartyid)).one()

        # one or both addr may be null, handle each case.
        if person.delivery_addr_id or person.mailing_addr_id:
            if person.delivery_addr_id and person.mailing_addr_id:
                expr = (CorpParty.delivery_addr_id == person.delivery_addr_id) | \
                    (CorpParty.mailing_addr_id == person.mailing_addr_id)
            elif person.delivery_addr_id:
                expr = (CorpParty.delivery_addr_id == person.delivery_addr_id)
            elif person.mailing_addr_id:
                expr = (CorpParty.mailing_addr_id == person.mailing_addr_id)

            same_addr = CorpParty.query.add_columns(
                Event.event_timestmp
            ).filter(expr).join(Event, Event.event_id == CorpParty.start_event_id)
        else:
            same_addr = []

        same_name_and_company = CorpParty.query.add_columns(
            Event.event_timestmp
        ).filter(
            CorpParty.first_nme.ilike(person.first_nme),
            CorpParty.last_nme.ilike(person.last_nme),
            CorpParty.corp_num.ilike(person.corp_num),
        ).join(Event, Event.event_id == CorpParty.start_event_id)


        return jsonify({
            'offices': offices,
            'same_addr': [{**s[0].as_dict(), **{'year':int(s[1].year)}} for s in same_addr if s[0].corp_party_id != int(corppartyid)],
            'same_name_and_company': [{**s[0].as_dict(), **{'year':int(s[1].year)}} for s in same_name_and_company if s[0].corp_party_id != int(corppartyid)],
        })

    return app


def _merge_corpparty_search_addr_fields(row):
    address = row.addr_line_1
    if row.addr_line_2:
        address += ", " + row.addr_line_2
    if row.addr_line_3:
        address += ", " + row.addr_line_3
    return address


def _is_addr_search(fields):
    return "addr_line_1" in fields or "postal_cd" in fields


def _add_additional_cols_to_search_results(args, row, result_dict):
    fields = args.getlist('field')
    additional_cols = args.get('additional_cols')
    if _is_addr_search(fields) or additional_cols == ADDITIONAL_COLS_ADDRESS:
        result_dict['addr'] = _merge_corpparty_search_addr_fields(row)
        result_dict['postal_cd'] = row.postal_cd
    elif additional_cols == ADDITIONAL_COLS_ACTIVE:
        result_dict['state_typ_cd'] = row.state_typ_cd


def _add_additional_cols_to_search_query(args, query):
    fields = args.getlist('field')
    additional_cols = args.get('additional_cols')
    if _is_addr_search(fields) or additional_cols == ADDITIONAL_COLS_ADDRESS:
        query = query.join(Address, CorpParty.mailing_addr_id == Address.addr_id)
        query = query.add_columns(
            Address.addr_line_1,
            Address.addr_line_2,
            Address.addr_line_3,
            Address.postal_cd)
    elif additional_cols == ADDITIONAL_COLS_ACTIVE:
        query = query.join(CorpState, CorpState.corp_num == CorpParty.corp_num)\
            .join(CorpOpState, CorpOpState.state_typ_cd == CorpState.state_typ_cd)
        query = query.add_columns(CorpOpState.state_typ_cd)

    return query


def _get_model_by_field(field_name):

    # return CorpParty
    # [cvo] for performance we only query this one above table for now.

    if field_name in ['first_nme', 'middle_nme', 'last_nme', 'appointment_dt', 'cessation_dt', 'corp_num',
                      'corp_party_id']:  # CorpParty fields
        return eval('CorpParty')
    # elif field_name in ['corp_num']: # Corporation fields
    #     return eval('Corporation')
    # elif field_name in ['corp_nme']: # CorpName fields
    #     return eval('CorpName')
    elif field_name in ['addr_line_1', 'addr_line_2', 'addr_line_3', 'postal_cd', 'city', 'province']:  # Address fields
        return eval('Address')


def _get_filter(field, operator, value):

    if field == 'any_nme':
        return (_get_filter('first_nme', operator, value)
            | _get_filter('middle_nme', operator, value)
            | _get_filter('last_nme', operator, value))

    if field == 'addr':
        # return _get_filter('first_nme', operator, value)
        return (_get_filter('addr_line_1', operator, value)
            | _get_filter('addr_line_2', operator, value)
            | _get_filter('addr_line_3', operator, value))

    model = _get_model_by_field(field)

    value = value.lower()
    if model:
        Field = getattr(model, field)
        # TODO: we should sanitize the values
        if operator == 'contains':
            return Field.ilike('%' + value + '%')
        elif operator == 'exact':
            return Field.ilike(value)
        elif operator == 'endswith':
            return Field.ilike('%' + value)
        elif operator == 'startswith':
            return Field.ilike(value + '%')
        elif operator == 'wildcard':
            return Field.ilike(value)
        else:
            raise Exception('invalid operator: {}'.format(operator))
    else:
        raise Exception('invalid field: {}'.format(field))


def _get_sort_field(field_name):

    model = _get_model_by_field(field_name)
    if model:
        return getattr(model, field_name)
    else:
        raise Exception('invalid sort field: {}'.format(field_name))


def _get_corporation_search_results(args):
    query = args.get("query")

    if not query:
        return "No search query was received", 400

    # TODO: move queries to model class.
    results = (
        Corporation.query
        .join(CorpName, Corporation.corp_num == CorpName.corp_num)
        # .join(CorpParty, Corporation.corp_num == CorpParty.corp_num)
        # .join(Office, Office.corp_num == Corporation.corp_num)
        # .join(Address, Office.mailing_addr_id == Address.addr_id)
        .with_entities(
            CorpName.corp_nme,
            Corporation.corp_num,
            # Corporation.transition_dt,
            # Address.addr_line_1,
            # Address.addr_line_2,
            # Address.addr_line_3,
            # Address.postal_cd,
            # Address.city,
            # Address.province,
        )
        # .filter(Office.end_event_id == None)
        # .filter(CorpName.end_event_id == None)
    )

    results = results.filter(
        (Corporation.corp_num == query) |
        (CorpName.corp_nme.ilike('%' + query + '%'))
        # (CorpParty.first_nme.contains(query)) |
        # (CorpParty.last_nme.contains(query)))
    )

    return results


def _get_corpparty_search_results(args):
    """
    Querystring parameters as follows:

    You may provide query=<string> for a simple search, OR any number of querystring triples such as

    field=ANY_NME|first_nme|last_nme|<any column name>
    &operator=exact|contains|startswith|endswith
    &value=<string>
    &sort_type=asc|desc
    &sort_value=ANY_NME|first_nme|last_nme|<any column name>
    &additional_cols=address|active|none

    For example, to get everyone who has any name that starts with 'Sky', or last name must be exactly 'Little', do:
    curl "http://localhost/person/search/?field=ANY_NME&operator=startswith&value=Sky&field=last_nme&operator=exact&value=Little&mode=ALL"
    """

    query = args.get("query")

    fields = args.getlist('field')
    operators = args.getlist('operator')
    values = args.getlist('value')
    mode = args.get('mode')
    sort_type = args.get('sort_type')
    sort_value = args.get('sort_value')

    if query and len(fields) > 0:
        raise Exception("use simple query or advanced. don't mix")

    # Only triples of clauses are allowed. So, the same number of fields, ops and values.
    if len(fields) != len(operators) or len(operators) != len(values):
        raise Exception("mismatched query param lengths: fields:{} operators:{} values:{}".format(
            len(fields),
            len(operators),
            len(values)))

    # Zip the lists, so ('last_nme', 'first_nme') , ('contains', 'exact'), ('Sky', 'Apple') => (('last_nme', 'contains', 'Sky'), ('first_nme', 'exact', 'Apple'))
    clauses = list(zip(fields, operators, values))

    # TODO: move queries to model class.
            # TODO: we no longer need this as we want to show all types.
            #.filter(CorpParty.party_typ_cd.in_(['FIO', 'DIR','OFF']))\

    results = (CorpParty.query
            # .filter(CorpParty.end_event_id == None)
            # .filter(CorpName.end_event_id == None)
            # .join(Corporation, Corporation.corp_num == CorpParty.corp_num)\
            # .join(CorpState, CorpState.corp_num == CorpParty.corp_num)\
            # .join(CorpOpState, CorpOpState.state_typ_cd == CorpState.state_typ_cd)\
            # .join(CorpName, Corporation.corp_num == CorpName.corp_num)\
            # .join(Address, CorpParty.mailing_addr_id == Address.addr_id)
            .add_columns(
                CorpParty.corp_party_id,
                CorpParty.first_nme,
                CorpParty.middle_nme,
                CorpParty.last_nme,
                CorpParty.appointment_dt,
                CorpParty.cessation_dt,
                CorpParty.corp_num,
                CorpParty.party_typ_cd,
                # Corporation.corp_num,
                # CorpName.corp_nme,
                # Address.addr_line_1,
                # Address.addr_line_2,
                # Address.addr_line_3,
                # Address.postal_cd,
                # Address.city,
                # Address.province,
                # CorpOpState.state_typ_cd,
                # CorpOpState.full_desc,
            ))

    results = _add_additional_cols_to_search_query(args, results)

    # Simple mode - return reasonable results for a single search string:
    if query:
        #results = results.filter((Corporation.corp_num == query) | (CorpParty.first_nme.contains(query)) | (CorpParty.last_nme.contains(query)))
        results = results.filter(CorpParty.first_nme.ilike(query) | CorpParty.last_nme.ilike(query) | CorpParty.middle_nme.ilike(query))
        # Advanced mode - return precise results for a set of clauses.
    elif clauses:

        # Determine if we will combine clauses with OR or AND. mode=ALL means we use AND. Default mode is OR
        if mode == 'ALL':
            def fn(accumulator, s):
                return accumulator & _get_filter(*s)
        else:
            def fn(accumulator, s):
                return accumulator | _get_filter(*s)

        # We use reduce here to join all the items in clauses with the & operator or the | operator.
        # Similar to if we did "|".join(clause), but calling the boolean operator instead.
        filter_grp = reduce(
            fn,
            clauses[1:],
            _get_filter(*clauses[0])
        )
        results = results.filter(filter_grp)

    # Sorting
    if sort_type is None:
        results = results.order_by(CorpParty.last_nme, CorpParty.corp_num)
    else:
        field = _get_sort_field(sort_value)

        if sort_type == 'desc':
            results = results.order_by(desc(field))
        else:
            results = results.order_by(field)

    # TODO: uncomment
    #raise Exception(results.statement.compile())
    return results


# if __name__ == '__main__':
#     app = create_app()
#     app.run(host='0.0.0.0')
